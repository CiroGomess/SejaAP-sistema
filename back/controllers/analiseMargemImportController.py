import re
from io import BytesIO
from decimal import Decimal, InvalidOperation
from flask import request, jsonify
from openpyxl import load_workbook
from config.db import get_connection

import pandas as pd
from datetime import datetime, date


# Reaproveita suas funções de salvar docs
from controllers.uploadDocsController import (
    _customer_id_exists,
    _save_file_for_customer,
)

DOCS_IMPORT_CATEGORIA_DEFAULT = "ANALISE_MARGEM_XLSX"


# -------------------------
# Helpers XLSX
# -------------------------
def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ç", "c")
    s = s.replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e")
    s = s.replace("í", "i")
    s = s.replace("ó", "o").replace("ô", "o")
    s = s.replace("ú", "u")
    s = re.sub(r"[^a-z0-9_ ]", "", s)
    s = s.replace(" ", "_")
    return s


def _to_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _to_decimal(v, default=None):
    """
    Aceita:
      - 6,70
      - 6.70
      - 1.234,56
      - 1234.56
    """
    if v is None or str(v).strip() == "":
        return default
    try:
        if isinstance(v, (int, float, Decimal)):
            return Decimal(str(v)).quantize(Decimal("0.01"))

        s = str(v).strip()

        # caso: tem "," como decimal -> remove "." milhar e troca "," por "."
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return default


def _try_calc_margem_bruta(payload: dict) -> None:
    """
    margem_bruta = custo + hora_homem + frete
                   + impostos (%) + comissão (%)
    """
    if payload.get("margem_bruta") is not None:
        return

    try:
        custo = Decimal(str(payload.get("custo", 0) or 0))
        hora = Decimal(str(payload.get("hora_homem", 0) or 0))
        frete = Decimal(str(payload.get("frete", 0) or 0))

        imposto_pct = Decimal(str(payload.get("imposto", 0) or 0)) / 100
        comissao_pct = Decimal(str(payload.get("comissao", 0) or 0)) / 100

        base = custo + hora + frete
        impostos = base * imposto_pct
        comissao = base * comissao_pct

        payload["margem_bruta"] = (base + impostos + comissao).quantize(Decimal("0.01"))
    except Exception:
        # se não der pra calcular, só deixa como None
        pass


def _read_xlsx_rows(file_bytes: bytes) -> list[dict]:
    """
    Lê a primeira sheet e retorna lista de dicts normalizados.
    Espera headers como na imagem:
      PRODUTO OU SERVIÇO | CUSTO | HORA HOMEM | IMPOSTO | MARGEM BRUTA | COMISSÃO | FRETE
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Headers (linha 1)
    headers = [_norm_header(_to_str(c.value)) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    # alias para tolerar pequenas variações
    # (ex.: "produto_ou_servico" ou "produto_ou_servico_")
    aliases = {
        "produto_ou_servico": ["produto_ou_servico", "produto_ou_servico_", "produto_servico"],
        "custo": ["custo"],
        "hora_homem": ["hora_homem", "hora_homens", "hora_homem_"],
        "imposto": ["imposto", "impostos"],
        "margem_bruta": ["margem_bruta", "margem", "margem_bruta_"],
        "comissao": ["comissao", "comissao_"],
        "frete": ["frete"],
    }

    def _find_key(possible_keys: list[str]) -> int | None:
        for k in possible_keys:
            if k in idx:
                return idx[k]
        return None

    map_idx = {k: _find_key(v) for k, v in aliases.items()}

    def get(row, key, default=None):
        i = map_idx.get(key)
        if i is None:
            return default
        if i >= len(row):
            return default
        return row[i]

    out = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if not any([c is not None and str(c).strip() != "" for c in row_cells]):
            continue

        item = {
            "produto_ou_servico": _to_str(get(row_cells, "produto_ou_servico")) or None,
            "custo": _to_decimal(get(row_cells, "custo"), default=None),
            "hora_homem": _to_decimal(get(row_cells, "hora_homem"), default=None),
            "imposto": _to_decimal(get(row_cells, "imposto"), default=None),       # % (ex 12.5)
            "margem_bruta": _to_decimal(get(row_cells, "margem_bruta"), default=None),
            "comissao": _to_decimal(get(row_cells, "comissao"), default=None),     # % (ex 5)
            "frete": _to_decimal(get(row_cells, "frete"), default=None),
        }

        # se o excel não trouxe margem_bruta, calcula
        _try_calc_margem_bruta(item)

        out.append(item)

    return out


# -------------------------
# Nova rota: upload + import
# -------------------------
def import_analise_margem_xlsx(current_user=None):
    file = request.files.get("file")
    id_cliente = request.form.get("id_cliente")
    categoria_doc = (request.form.get("categoria") or "Analise de Margem").strip()

    if not file or not id_cliente:
        return jsonify({"error": "Arquivo e id_cliente são obrigatórios"}), 400

    try:
        user_id = int(id_cliente)
    except:
        return jsonify({"error": "id_cliente inválido"}), 400

    file_bytes = file.read()
    
    try:
        file.seek(0)
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(BytesIO(file_bytes), sep=',')
            except:
                df = pd.read_csv(BytesIO(file_bytes), sep=';')
        else:
            df = pd.read_excel(BytesIO(file_bytes))
        
        parsed_data = df.to_dict(orient="records")
    except Exception as e:
        return jsonify({"error": "Erro ao ler arquivo", "details": str(e)}), 400

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Salva histórico do arquivo
        file.stream = BytesIO(file_bytes)
        file.stream.seek(0)
        caminho_arquivo, _ = _save_file_for_customer(cur, user_id, categoria_doc, file)

        cur.execute(
            "INSERT INTO public.docs_clientes (id_cliente, caminho_arquivo, categoria) VALUES (%s, %s, %s) RETURNING id",
            (user_id, caminho_arquivo, categoria_doc)
        )
        doc_id = cur.fetchone()[0]

        insert_sql = """
            INSERT INTO public.analise_margem
            (user_id, produto_ou_servico, custo, hora_homem, imposto, margem_bruta, comissao, frete, data_registro)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s);
        """

        values_to_insert = []
        for item in parsed_data:
            # Inicializa variáveis
            data_raw = None
            prod_serv = None
            custo = 0
            hora_homem = 0
            imposto = 0
            margem_bruta = 0
            comissao = 0
            frete = 0

            # BUSCA DINÂMICA PARA TODOS OS CAMPOS
            for key, value in item.items():
                k_norm = str(key).strip().lower().replace(" ", "").replace("-", "").replace("_", "").replace("ç", "c").replace("ã", "a")
                
                if "registro" in k_norm or "data" in k_norm:
                    data_raw = value
                elif "produto" in k_norm or "servico" in k_norm or "desc" in k_norm:
                    prod_serv = value
                elif "custo" in k_norm:
                    custo = value
                elif "hora" in k_norm:
                    hora_homem = value
                elif "imposto" in k_norm:
                    imposto = value
                elif "margem" in k_norm:
                    margem_bruta = value
                elif "comissao" in k_norm:
                    comissao = value
                elif "frete" in k_norm:
                    frete = value

            if not prod_serv: continue

            # Tratamento da Data
            data_formatada = None
            if data_raw and str(data_raw).lower() != 'nan':
                if isinstance(data_raw, (date, datetime)):
                    data_formatada = data_raw if isinstance(data_raw, date) else data_raw.date()
                elif isinstance(data_raw, str) and data_raw.strip():
                    try:
                        data_formatada = datetime.strptime(data_raw.strip().split(" ")[0], "%d/%m/%Y").date()
                    except:
                        try: data_formatada = date.fromisoformat(data_raw.strip().split(" ")[0])
                        except: data_formatada = None

            values_to_insert.append((
                user_id,
                str(prod_serv),
                float(custo) if custo and str(custo).lower() != 'nan' else 0,
                float(hora_homem) if hora_homem and str(hora_homem).lower() != 'nan' else 0,
                float(imposto) if imposto and str(imposto).lower() != 'nan' else 0,
                float(margem_bruta) if margem_bruta and str(margem_bruta).lower() != 'nan' else 0,
                float(comissao) if comissao and str(comissao).lower() != 'nan' else 0,
                float(frete) if frete and str(frete).lower() != 'nan' else 0,
                data_formatada
            ))

        if values_to_insert:
            cur.executemany(insert_sql, values_to_insert)
            conn.commit()

        return jsonify({"message": "Importado com sucesso", "doc_id": doc_id, "rows": len(values_to_insert)}), 201

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if conn: conn.close()
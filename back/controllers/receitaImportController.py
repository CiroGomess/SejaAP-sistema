import re
from io import BytesIO
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from flask import request, jsonify
from openpyxl import load_workbook

from config.db import get_connection
from utils.helpers import generate_secure_id

# Reaproveita sua função de salvar docs
from controllers.uploadDocsController import _save_file_for_customer

DOCS_IMPORT_CATEGORIA_DEFAULT = "RECEITAS_XLSX"


# -------------------------
# Helpers Cliente
# -------------------------
def _customer_id_exists(cur, customer_id: str) -> bool:
    cur.execute(
        """
        SELECT 1
        FROM public.clientes
        WHERE id = %s
        LIMIT 1;
        """,
        (customer_id,),
    )
    return cur.fetchone() is not None


def _get_customer_display_name(cur, id_cliente: str) -> str | None:
    """
    Retorna um nome amigável do cliente:
      1) first_name + last_name
      2) company_name
      3) code
      4) id_cliente
    """
    cur.execute(
        """
        SELECT
            COALESCE(NULLIF(first_name, ''), NULL)   AS first_name,
            COALESCE(NULLIF(last_name, ''), NULL)    AS last_name,
            COALESCE(NULLIF(company_name, ''), NULL) AS company_name,
            COALESCE(NULLIF(code, ''), NULL)         AS code
        FROM public.clientes
        WHERE id = %s
        LIMIT 1;
        """,
        (id_cliente,),
    )
    r = cur.fetchone()
    if not r:
        return None

    first_name, last_name, company_name, code = r
    full = " ".join([x for x in [first_name, last_name] if x]).strip()

    if full:
        return full
    if company_name:
        return str(company_name)
    if code:
        return str(code)
    return str(id_cliente)


# -------------------------
# Helpers XLSX
# -------------------------
def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ç", "c")
    s = s.replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e")
    s = s.replace("í", "i")
    s = s.replace("ó", "o").replace("ô", "o")
    s = s.replace("ú", "u")
    s = re.sub(r"[^a-z0-9_ ]", "", s)
    s = s.replace(" ", "_")
    return s


def _to_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v, default=0):
    if v is None or str(v).strip() == "":
        return default
    try:
        return int(float(str(v).replace(",", ".")))
    except Exception:
        return default


def _to_decimal(v, default=None):
    """
    Aceita:
      - 6,70
      - 6.70
      - 1.234,56
      - 1234.56
    """
    if v is None or str(v).strip() == "":
        return default
    try:
        if isinstance(v, (int, float, Decimal)):
            return Decimal(str(v)).quantize(Decimal("0.01"))

        s = str(v).strip()

        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return default


def _to_date_iso(v) -> str | None:
    """
    Retorna YYYY-MM-DD ou None.
    """
    if v is None or str(v).strip() == "":
        return None

    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()

    s = str(v).strip()

    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s

    m = re.match(r"^(\d{2})/(\d{2})/(\d{2}|\d{4})$", s)
    if m:
        dd, mm, yy = m.group(1), m.group(2), m.group(3)
        if len(yy) == 2:
            yy = "20" + yy
        try:
            return date(int(yy), int(mm), int(dd)).isoformat()
        except Exception:
            return None

    return None


def _read_xlsx_rows(file_bytes: bytes) -> list[dict]:
    """
    Lê a primeira sheet e retorna uma lista de dicts normalizados.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(_norm_header(_to_str(cell.value)))

    idx = {h: i for i, h in enumerate(headers) if h}

    def get(row, key, default=None):
        i = idx.get(key)
        if i is None:
            return default
        if i >= len(row):
            return default
        return row[i]

    out = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if not any([c is not None and str(c).strip() != "" for c in row_cells]):
            continue

        item = {
            "n_orcamento": _to_str(get(row_cells, "n_orcamento")),
            "nome_fantasia": _to_str(get(row_cells, "nome_fantasia")),
            "data_emissao": _to_date_iso(get(row_cells, "data_emissao")),
            "data_vencimento": _to_date_iso(get(row_cells, "vencimento")),
            "cod_produto": _to_str(get(row_cells, "cod_produto")),
            "descricao": _to_str(get(row_cells, "descricao")),
            "qtd": _to_int(get(row_cells, "qtd"), default=0),
            "unitario": _to_decimal(get(row_cells, "unitario"), default=Decimal("0.00")),
            "total": _to_decimal(get(row_cells, "total"), default=None),
            "unidade_filial": _to_str(get(row_cells, "unidade_filial")) or None,
        }

        if item["total"] is None:
            try:
                item["total"] = (
                    Decimal(str(item["qtd"])) * Decimal(str(item["unitario"]))
                ).quantize(Decimal("0.01"))
            except Exception:
                item["total"] = None

        out.append(item)

    return out


# -------------------------
# Nova rota: upload + import
# -------------------------
def import_receitas_xlsx(current_user=None):
    """
    POST /receitas/import-xlsx
    multipart/form-data:
      - file (xlsx)
      - id_cliente (clientes.id) -> vira user_id em receita_user
      - (opcional) categoria
      - (opcional) projeto, centro_de_resultado
    """
    file = request.files.get("file")
    id_cliente = request.form.get("id_cliente")
    categoria = request.form.get("categoria") or DOCS_IMPORT_CATEGORIA_DEFAULT

    projeto = (request.form.get("projeto") or "IMPORT_XLSX").strip()
    centro_de_resultado = (request.form.get("centro_de_resultado") or "IMPORT_XLSX").strip()

    if not file:
        return jsonify({"error": "Missing file", "details": "field 'file' is required"}), 400
    if not id_cliente:
        return jsonify({"error": "Missing id_cliente"}), 400

    user_id = str(id_cliente).strip()
    if not user_id:
        return jsonify({"error": "Invalid id_cliente", "details": "id_cliente is required"}), 400

    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "Empty file"}), 400

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        if not _customer_id_exists(cur, user_id):
            return jsonify(
                {"error": "Invalid id_cliente", "details": "id_cliente must exist in public.clientes.id"}
            ), 400

        customer_name = _get_customer_display_name(cur, user_id)
        if not customer_name:
            return jsonify({"error": "Invalid id_cliente", "details": "Customer not found"}), 400

        # 1) salva arquivo físico + registra em docs_clientes
        file.stream = BytesIO(file_bytes)
        file.stream.seek(0)

        caminho_arquivo, err = _save_file_for_customer(cur, user_id, categoria, file)
        if err:
            return err

        novo_doc_id = generate_secure_id()

        cur.execute(
            """
            INSERT INTO public.docs_clientes (id, id_cliente, caminho_arquivo, categoria)
            VALUES (%s, %s, %s, %s)
            RETURNING id, id_cliente, caminho_arquivo, categoria, data_envio;
            """,
            (novo_doc_id, user_id, caminho_arquivo, categoria),
        )
        doc_row = cur.fetchone()

        # 2) parse XLSX
        mapped = _read_xlsx_rows(file_bytes)
        if not mapped:
            conn.rollback()
            return jsonify({"error": "No rows found", "details": "XLSX has no data rows"}), 400

        # 3) monta payloads e insere em lote
        receitas_payloads = []
        for it in mapped:
            nome_item = it["descricao"]
            if it["cod_produto"]:
                nome_item = f"[{it['cod_produto']}] {nome_item}".strip()

            payload = {
                "id": generate_secure_id(),
                "user_id": user_id,
                "numero_orcamento": it["n_orcamento"] or None,
                "nome_cliente": it["nome_fantasia"] or customer_name,
                "data_emissao": it["data_emissao"],
                "data_vencimento": it["data_vencimento"],
                "produto_ou_servico": "PRODUTO",
                "nome_produto_ou_servico": nome_item or None,
                "quantidade": it["qtd"],
                "valor_unitario": str(it["unitario"]) if it["unitario"] is not None else None,
                "valor_total": str(it["total"]) if it["total"] is not None else None,
                "unidade_filial": it.get("unidade_filial"),
                "projeto": projeto or None,
                "centro_de_resultado": centro_de_resultado or None,
            }
            receitas_payloads.append(payload)

        insert_sql = """
            INSERT INTO public.receita_user
            (id, user_id, numero_orcamento, nome_cliente, data_emissao, data_vencimento,
             produto_ou_servico, nome_produto_ou_servico, quantidade,
             valor_unitario, valor_total, unidade_filial, projeto, centro_de_resultado)
            VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """

        values = []
        for p in receitas_payloads:
            values.append(
                (
                    p["id"],
                    p["user_id"],
                    p["numero_orcamento"],
                    p["nome_cliente"],
                    p["data_emissao"],
                    p["data_vencimento"],
                    p["produto_ou_servico"],
                    p["nome_produto_ou_servico"],
                    p["quantidade"],
                    p["valor_unitario"],
                    p["valor_total"],
                    p["unidade_filial"],
                    p["projeto"],
                    p["centro_de_resultado"],
                )
            )

        cur.executemany(insert_sql, values)
        conn.commit()

        doc_json = {
            "id": doc_row[0],
            "id_cliente": doc_row[1],
            "caminho_arquivo": doc_row[2],
            "categoria": doc_row[3],
            "data_envio": str(doc_row[4]) if doc_row[4] else None,
        }

        return jsonify(
            {
                "message": "XLSX imported and receitas created",
                "doc": doc_json,
                "customer_name": customer_name,
                "rows_mapped": len(mapped),
                "rows_inserted": len(receitas_payloads),
                "mapped_preview": mapped[:5],
            }
        ), 201

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"error": "import failed", "details": str(e)}), 500
    finally:
        if conn:
            conn.close()